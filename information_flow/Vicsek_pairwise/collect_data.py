# import h5py
import numpy as np
import os
import openpyxl
import sys
sys.path.insert(0, sys.path[0]+"\\..\\")

from h5py_process import *




if __name__ == "__main__":
    path_of_this_file = os.path.split(__file__)[0]
    os.chdir(path_of_this_file)


    # enter the path where the data is
    os.chdir("Data")

    

    
    #to run this code you must know the list of eta and wlf
    
    eta_list = np.arange(0, 2.01, .2)*np.pi
    wFL_list = [1.0, 1.1, 1.2, 1.3, 1.4, 1.5, 1.7, 1.9, 2.0,3.0,4.0,5.0,7.0,9.0,10.0,15,20,25,30,40,50,60,70,80,90,100] # argv[6]


    # parameter initialization
    number_of_bins = 8

    informationList = {
        "A_at":16,
        "A_ta":16,


        "total_mutul_information":16,
        "TDMI": 16,
        "TE":16,

        "IMI_unique0":16, # the strategy of imi is not consistant
        "IMI_redundance0":16,
        "IMI_synergy0":16,

        "IMI_unique1":16, # the strategy of imi is not consistant
        "IMI_redundance1":16,
        "IMI_synergy1":16,
    
        "I_min_unique0":16,
        "I_min_unique1":16,
        "I_min_redundance":16,
        "I_min_synergy":16,
        
        "surd_unique0":16,
        "surd_unique1":16,
        "surd_redundance":16,
        "surd_synergy":16,
        "surd_information_leak":16,

        # normalized by mutual information
        "nTDMI": 16,
        "nTE":16,
        
        "nIMI_unique0":16, # the strategy of imi is not consistant
        "nIMI_redundance0":16,
        "nIMI_synergy0":16,

        "nIMI_unique1":16, # the strategy of imi is not consistant
        "nIMI_redundance1":16,
        "nIMI_synergy1":16,
    
        "nI_min_unique0":16,
        "nI_min_unique1":16,
        "nI_min_redundance":16,
        "nI_min_synergy":16,

        "nsurd_unique0":16,
        "nsurd_unique1":16,
        "nsurd_redundance":16,
        "nsurd_synergy":16,
        }

    wb = openpyxl.Workbook()
    
    sheet = wb.create_sheet("information", 0)
    quantity_list = []



    eta_list = np.round(eta_list,8)
    for key in informationList:
        quantity_list.append(key)
        quantity_index= list(quantity_list).index(key)
        sheet.cell(row=1,column=1+quantity_index*(len(eta_list)+1)).value = key + "_eta(col)_wFL(row)"


        for i in range(np.shape(eta_list)[0]):
            sheet.cell(row=1,column=2+i).value = str(round(eta_list[i]/np.pi,2))
        for i in range(np.shape(wFL_list)[0]):
            sheet.cell(row=2+i,column=1).value = str(round(wFL_list[i],2))

    
    for folder in os.listdir("."): # folder is the name of each folder
        if folder[-5:] == '.xlsx' or folder[-4:] == '.txt':
            continue
        print(folder)
        os.chdir(folder)

        f = H5PY_Processor("Data-pairwise_inf-"+str(number_of_bins)+".hdf5","r")

        number_of_particles  = f.f["number_of_particles"][:][0]
        number_of_influencers = f.f["number_of_influencers"][:][0]
        wFL = f.f["wLF"][:][0] # the difference in name is the convention
        eta = f.f["eta"][:][0]
        dataNames = f.f.keys()
        

        for key in dataNames:
            if key in informationList:
                informationList[key] = f.f[key][:][0]
        # write the data
        for keys in dataNames:
            if keys in informationList:
                quantity_index= list(quantity_list).index(keys)
                eta_index= list(eta_list).index(round(eta,8))
                wFL_index= list(wFL_list).index(wFL)

                sheet.cell(row= 2 + wFL_index, column= 2 + eta_index+quantity_index*(len(eta_list)+1)).value = informationList[keys]#和画图的格式一样
        f.close()

        os.chdir("..")

    wb.save('inf_pairwise_'+str(number_of_bins)+'.xlsx')




    