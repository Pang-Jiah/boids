# Collect the order parameters and exhibit them in an excel file

import numpy as np
import os
import openpyxl

import sys
from order_parameters import *
sys.path.insert(0, sys.path[0]+"\\..\\")
from Vicsek import *
from h5py_process import *

def average_and_ssd(start_point:int, stop_point:int, data:np.ndarray):
    '''
    calculate the v_a in an interval of time

    Return
    average: the average 
    standard_deviation: sample standard deviation
    '''
    data = np.float64(data)
    data  = data.reshape((1,-1))
    average = np.average(data[0][start_point:stop_point])
    sample_standard_deviation = np.var(data[0][start_point:stop_point], ddof = 1)
    return average, sample_standard_deviation




if __name__ == "__main__":
    mypath = os.path.split(__file__)[0]
    os.chdir(mypath)


    # enter the path where your data is 
    os.chdir("Data_40")

    # initial the data
    Va_list = [] 
    Va_list_ssd = []

    A_pant = []
    A_pant_ssd = []

    A_pant_step = []
    A_pant_step_ssd = []

    A_apnt_step = []
    A_apnt_step_ssd = []

    A_tapn_list = []
    A_tapn_ssd_list = []

    A_tpn_list = []
    A_tpn_ssd_list = []

    eta_all = []


    for folder in os.listdir("."): # folder is the name of each folder
        if folder[-5:] == '.xlsx' or folder[-4:] == '.txt':
            continue
        print(folder)
        os.chdir(folder)

        # Read the data
        f = H5PY_Processor("Data-Vicsek_collective.hdf5","r")
        va = f.f["v_a"][:]
        a_pan_step  = f.f["A_pan_step"][:]
        a_apn_step  = f.f["A_apn_step"][:]
        a_tapn  = f.f["A_tapn"][:][0]
        a_tapn_ssd  = f.f["A_tapn_ssd"][:][0]
        a_tpna  = f.f["A_tpna"][:][0]
        a_tpn_ssd  = f.f["A_tpn_ssd"][:][0]
        groupParameter = f.f["parameters"]
        eta            = groupParameter["noise_strength"][:][0]
        number_of_steps         = groupParameter["number_of_steps"][:][0]
        number_of_particles     = groupParameter["number_of_particles"][:][0]
        size_of_arena           = groupParameter["size_of_arena"][:][0]
        f.close()




        os.chdir("..")
        start_point =1000 # this parameter should be consist with the parameter in Vicsek_collective.py
        stop_point =number_of_steps 

        Va_average, Va_ssd = average_and_ssd(start_point=start_point, stop_point=stop_point, data=va)
        Va_list.append(Va_average)
        # Va_list_ssd.append(Va_ssd)

        pant_step_average, pant_step_ssd = average_and_ssd(start_point=start_point, stop_point=stop_point, data=a_pan_step)
        A_pant_step.append(pant_step_average)
        # A_pant_step_ssd.append(pant_step_ssd)

        
        apnt_step_average, apnt_step_ssd = average_and_ssd(start_point=start_point, stop_point=stop_point, data=a_apn_step)
        A_apnt_step.append(apnt_step_average)
        # A_apnt_step_ssd.append(apnt_step_ssd)


        A_tapn_list.append(a_tapn)
        # A_tapn_ssd_list.append(a_tapn_ssd)

        A_tpn_list.append(a_tpna)
        # A_tpn_ssd_list.append(a_tpn_ssd)



        # rho_all.append(number_of_particles/size_of_arena**2) # for different density rho
        eta_all.append(eta)



    # print(np.array(eta_all)/3.14)
    eta_list  = np.arange(0.0, 2.01, .05)*np.pi
    eta_list = np.round(eta_list,3)



    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("order_parameters", 0)
    sheet.cell(row=1,column=1).value = "eta (pi)"
    # sheet.cell(row=1,column=1).value = "density" # for different density rho
    sheet.cell(row=1,column=2).value = "Va"
    sheet.cell(row=1,column=3).value = "A_pant_step"
    sheet.cell(row=1,column=4).value = "A_apnt_step"
    sheet.cell(row=1,column=5).value = "A_tapn"
    sheet.cell(row=1,column=6).value = "A_tpn"

    # sheet.cell(row=1,column=7).value = "Va_ssd"
    # sheet.cell(row=1,column=8).value = "A_pant_step_ssd"
    # sheet.cell(row=1,column=9).value = "A_apnt_step_ssd"
    # sheet.cell(row=1,column=10).value = "A_tapn_ssd"
    # sheet.cell(row=1,column=11).value = "A_tpn_ssd"

    stride = 1

    # store the data
    for i in range(np.shape(eta_all)[0]):

        eta_index= list(eta_list).index(round(eta_all[i], 3))
        sheet.cell(row=stride*(eta_index)+2,column=1).value = np.round(eta_all[i]/np.pi, decimals=3)
        # sheet.cell(row=stride*(i)+2,column=1).value = "{:.5f}".format(eta_all[i]/np.pi) # for different density rho
        
        sheet.cell(row=stride*(eta_index)+2,column=2).value = Va_list[i]
        sheet.cell(row=stride*(eta_index)+2,column=3).value = A_pant_step[i]
        sheet.cell(row=stride*(eta_index)+2,column=4).value = A_apnt_step[i]
        sheet.cell(row=stride*(eta_index)+2,column=5).value = A_tapn_list[i]
        sheet.cell(row=stride*(eta_index)+2,column=6).value = A_tpn_list[i]


        # sheet.cell(row=stride*(eta_index)+2,column=7).value = Va_list_ssd[i]
        # sheet.cell(row=stride*(eta_index)+2,column=8).value = A_pant_step_ssd[i]
        # sheet.cell(row=stride*(eta_index)+2,column=9).value = A_apnt_step_ssd[i]
        # sheet.cell(row=stride*(eta_index)+2,column=10).value = A_tapn_ssd_list[i]
        # sheet.cell(row=stride*(eta_index)+2,column=11).value = A_tpn_ssd_list[i]

    wb.save('order_parameters.xlsx')




    