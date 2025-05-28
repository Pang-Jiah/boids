# Collect the values of the information-theoretic quantities 
import numpy as np
import os
import openpyxl
import sys

sys.path.insert(0, sys.path[0]+"\\..\\")
from h5py_process import *

if __name__ == "__main__":
    path_of_this_file = os.path.split(__file__)[0]
    os.chdir(path_of_this_file)
    
    # enter the folder where the data is stored
    os.chdir("Data_40")

    number_of_bins = 8

    #to run this code you must know the list of eta
    eta_list  = np.arange(0.0, 2.01, .05)*np.pi
    eta_list = np.round(eta_list,3)
    
    number_of_data_collection = 100
    informationList = {
                    "total_mutul_information":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "TDMI": np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "TE":np.ones(number_of_data_collection, dtype=np.float16)*16,

                    "IMI_unique0":np.ones(number_of_data_collection, dtype=np.float16)*16, # the strategy of imi is not consistant
                    "IMI_redundance0":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "IMI_synergy0":np.ones(number_of_data_collection, dtype=np.float16)*16,

                    "IMI_unique1":np.ones(number_of_data_collection, dtype=np.float16)*16, # the strategy of imi is not consistant
                    "IMI_redundance1":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "IMI_synergy1":np.ones(number_of_data_collection, dtype=np.float16)*16,
                
                
                    "I_min_unique0":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "I_min_unique1":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "I_min_redundance":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "I_min_synergy":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    
                    "surd_unique0":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "surd_unique1":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "surd_redundance":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "surd_synergy":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "surd_information_leak":np.ones(number_of_data_collection, dtype=np.float16)*16,

                    # normalized by mutual information
                    "nTDMI": np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nTE":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    
                    "nIMI_unique0":np.ones(number_of_data_collection, dtype=np.float16)*16, # the strategy of imi is not consistant
                    "nIMI_redundance0":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nIMI_synergy0":np.ones(number_of_data_collection, dtype=np.float16)*16,

                    "nIMI_unique1":np.ones(number_of_data_collection, dtype=np.float16)*16, # the strategy of imi is not consistant
                    "nIMI_redundance1":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nIMI_synergy1":np.ones(number_of_data_collection, dtype=np.float16)*16,
                
                    "nI_min_unique0":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nI_min_unique1":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nI_min_redundance":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nI_min_synergy":np.ones(number_of_data_collection, dtype=np.float16)*16,

                    "nsurd_unique0":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nsurd_unique1":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nsurd_redundance":np.ones(number_of_data_collection, dtype=np.float16)*16,
                    "nsurd_synergy":np.ones(number_of_data_collection, dtype=np.float16)*16,
   
                    }
    wb = openpyxl.Workbook()
    
    sheet = wb.create_sheet("information", 0)
    sheet.cell(row=1,column=1).value = "eta"

    for i in range(np.shape(eta_list)[0]):
        sheet.cell(row=2+i,column=1).value = (round(eta_list[i], 2))

    quantity_list = []
    for key in informationList:
        quantity_list.append(key)
        quantity_index= list(quantity_list).index(key)
        sheet.cell(row=1,column=(quantity_index)+2).value = key
        sheet.cell(row=1,column=(quantity_index)+50).value = str("ssd_")+key
    
    for folder in os.listdir("."): # folder is the name of each folder
        if folder[-5:] == '.xlsx' or folder[-4:] == '.txt' or folder[-5:] == '.opju':
            continue
        print(folder)
        os.chdir(folder)
        
        f = H5PY_Processor("Data-collective_inf-"+str(number_of_bins)+".hdf5","r")
        # number_of_particles  = f.f["number_of_particles"][:][0]
        # number_of_influencers = f.f["number_of_influencers"][:][0]
        # wFL = f.f["wLF"][:][0] # 
        # number_of_data_collection = f.f["number_of_data_collection"][:][0]
        number_of_data_collection_r = f.f["number_of_data_collection_r"][:][0]
        eta = f.f["eta"][:][0]
      
        dataNames = f.f.keys()

        for key in dataNames:
            if key in informationList:
                informationList[key] = f.f[key][:]

        for keys in dataNames:
            eta_index= list(eta_list).index(round(eta, 3))
            if keys in informationList:
                quantity_index= list(quantity_list).index(keys)
                sheet.cell(row=2+eta_index, column=(quantity_index)+2).value   = np.average(informationList[keys][:number_of_data_collection_r])#和画图的格式一样
                sheet.cell(row=2+eta_index, column=(quantity_index)+len(informationList)+2).value  = np.var(informationList[keys][:number_of_data_collection_r],ddof=1)#和画图的格式一样
        f.close()
        os.chdir("..")
        
    wb.save('inf_collective_'+str(number_of_bins)+'.xlsx')




    